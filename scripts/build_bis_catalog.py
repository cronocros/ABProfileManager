#!/usr/bin/env python3
"""
Build ABProfileManager/Data/BISCatalog.lua from refreshed Wowhead BIS seed tables,
the DOC seed files, and localized Retail DB2 item metadata.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import DefaultDict, Dict, Iterable, List, Optional, Tuple

import requests


REPO_ROOT = Path(__file__).resolve().parents[1]
OVERALL_FILE = REPO_ROOT / "ABProfileManager" / "Data" / "BISData_Method.lua"
FALLBACK_FILE = REPO_ROOT / "ABProfileManager" / "Data" / "BISData.lua"
DOC_FINAL_FILE = REPO_ROOT / "DOC" / "wow_midnight_s1_mplus_bis_final.md"
DOC_COMPANION_FILE = REPO_ROOT / "DOC" / "wow_midnight_s1_mplus_bis_korean_companion.md"
ADDON_DB_FILE = REPO_ROOT / "DOC" / "MidnightS1_MPlus_Addon_DB_v1.0.lua"
TARGET_FILE = REPO_ROOT / "ABProfileManager" / "Data" / "BISCatalog.lua"
WAGO_DB2_INDEX_URL = "https://wago.tools/db2"
WAGO_DB2_URL_TEMPLATE = "https://wago.tools/db2/{table}/csv?build={build}{locale_param}"
RETAIL_DB2_PREFIX = "12.0.1."
REQUEST_TIMEOUT_SECONDS = 120

DOC_SECTION_SPEC_IDS = {
    1: 269, 2: 268, 3: 270, 4: 250, 5: 581, 6: 104, 7: 66, 8: 73, 9: 577, 10: 103,
    11: 263, 12: 105, 13: 65, 14: 256, 15: 257, 16: 62, 17: 63, 18: 64, 19: 262,
    20: 251, 21: 70, 22: 258, 23: 1467, 24: 1468, 25: 72, 26: 253, 27: 255, 28: 252,
    29: 102, 30: 1382, 31: 1473, 32: 254, 33: 259, 34: 260, 35: 261, 36: 264,
    37: 265, 38: 266, 39: 267, 40: 71,
}

SPEC_NAMES = {
    62: "Arcane Mage", 63: "Fire Mage", 64: "Frost Mage", 65: "Holy Paladin",
    66: "Protection Paladin", 70: "Retribution Paladin", 71: "Arms Warrior",
    72: "Fury Warrior", 73: "Protection Warrior", 102: "Balance Druid",
    103: "Feral Druid", 104: "Guardian Druid", 105: "Restoration Druid",
    250: "Blood Death Knight", 251: "Frost Death Knight", 252: "Unholy Death Knight",
    253: "Beast Mastery Hunter", 254: "Marksmanship Hunter", 255: "Survival Hunter",
    256: "Discipline Priest", 257: "Holy Priest", 258: "Shadow Priest",
    259: "Assassination Rogue", 260: "Outlaw Rogue", 261: "Subtlety Rogue",
    262: "Elemental Shaman", 263: "Enhancement Shaman", 264: "Restoration Shaman",
    265: "Affliction Warlock", 266: "Demonology Warlock", 267: "Destruction Warlock",
    268: "Brewmaster Monk", 269: "Windwalker Monk", 270: "Mistweaver Monk",
    577: "Havoc Demon Hunter", 581: "Vengeance Demon Hunter", 1382: "Devourer Demon Hunter",
    1467: "Devastation Evoker", 1468: "Preservation Evoker", 1473: "Augmentation Evoker",
}

SOURCE_GROUP_ORDER = {"mythicplus": 0, "raid": 1, "crafted": 2, "tier": 3}
SOURCE_WEIGHT = {"overall": 100, "doc": 200, "fallback": 300}
NOTE_WEIGHT = {"BIS": 0, "대체재": 10, "2순위": 20, "3순위": 30}

SPEC_KEY_TO_ID = {
    "MAGE_ARCANE": 62,
    "MAGE_FIRE": 63,
    "MAGE_FROST": 64,
    "PALADIN_HOLY": 65,
    "PALADIN_PROTECTION": 66,
    "PALADIN_RETRIBUTION": 70,
    "WARRIOR_ARMS": 71,
    "WARRIOR_FURY": 72,
    "WARRIOR_PROTECTION": 73,
    "DRUID_BALANCE": 102,
    "DRUID_FERAL": 103,
    "DRUID_GUARDIAN": 104,
    "DRUID_RESTORATION": 105,
    "DEATHKNIGHT_BLOOD": 250,
    "DEATHKNIGHT_FROST": 251,
    "DEATHKNIGHT_UNHOLY": 252,
    "HUNTER_BEAST_MASTERY": 253,
    "HUNTER_MARKSMANSHIP": 254,
    "HUNTER_SURVIVAL": 255,
    "PRIEST_DISCIPLINE": 256,
    "PRIEST_HOLY": 257,
    "PRIEST_SHADOW": 258,
    "ROGUE_ASSASSINATION": 259,
    "ROGUE_OUTLAW": 260,
    "ROGUE_SUBTLETY": 261,
    "SHAMAN_ELEMENTAL": 262,
    "SHAMAN_ENHANCEMENT": 263,
    "SHAMAN_RESTORATION": 264,
    "WARLOCK_AFFLICTION": 265,
    "WARLOCK_DEMONOLOGY": 266,
    "WARLOCK_DESTRUCTION": 267,
    "MONK_BREWMASTER": 268,
    "MONK_WINDWALKER": 269,
    "MONK_MISTWEAVER": 270,
    "DEMONHUNTER_HAVOC": 577,
    "DEMONHUNTER_VENGEANCE": 581,
    "DEMONHUNTER_DEVOURER": 1382,
    "EVOKER_DEVASTATION": 1467,
    "EVOKER_PRESERVATION": 1468,
    "EVOKER_AUGMENTATION": 1473,
}

ADDON_DB_SPEC_KEY_ALIASES = {
    "DRUID_RESTO": "DRUID_RESTORATION",
    "HUNTER_BM": "HUNTER_BEAST_MASTERY",
    "HUNTER_MARKSMAN": "HUNTER_MARKSMANSHIP",
    "PRIEST_DISC": "PRIEST_DISCIPLINE",
}

SPEC_ID_TO_CLASS_TOKEN = {
    62: "MAGE", 63: "MAGE", 64: "MAGE",
    65: "PALADIN", 66: "PALADIN", 70: "PALADIN",
    71: "WARRIOR", 72: "WARRIOR", 73: "WARRIOR",
    102: "DRUID", 103: "DRUID", 104: "DRUID", 105: "DRUID",
    250: "DEATHKNIGHT", 251: "DEATHKNIGHT", 252: "DEATHKNIGHT",
    253: "HUNTER", 254: "HUNTER", 255: "HUNTER",
    256: "PRIEST", 257: "PRIEST", 258: "PRIEST",
    259: "ROGUE", 260: "ROGUE", 261: "ROGUE",
    262: "SHAMAN", 263: "SHAMAN", 264: "SHAMAN",
    265: "WARLOCK", 266: "WARLOCK", 267: "WARLOCK",
    268: "MONK", 269: "MONK", 270: "MONK",
    577: "DEMONHUNTER", 581: "DEMONHUNTER", 1382: "DEMONHUNTER",
    1467: "EVOKER", 1468: "EVOKER", 1473: "EVOKER",
}

CLASS_WEAPON_TYPES = {
    "DEATHKNIGHT": {"AXE_1H", "MACE_1H", "SWORD_1H", "AXE_2H", "MACE_2H", "SWORD_2H", "POLEARM"},
    "DEMONHUNTER": {"AXE_1H", "FIST", "SWORD_1H", "WARGLAIVE"},
    "DRUID": {"DAGGER", "FIST", "MACE_1H", "MACE_2H", "POLEARM", "STAFF"},
    "EVOKER": {"DAGGER", "FIST", "MACE_1H", "STAFF", "SWORD_1H", "OFF_HAND"},
    "HUNTER": {"BOW", "CROSSBOW", "GUN", "POLEARM", "STAFF"},
    "MAGE": {"DAGGER", "STAFF", "SWORD_1H", "WAND", "OFF_HAND"},
    "MONK": {"AXE_1H", "FIST", "MACE_1H", "POLEARM", "STAFF", "SWORD_1H"},
    "PALADIN": {"AXE_1H", "MACE_1H", "SWORD_1H", "AXE_2H", "MACE_2H", "POLEARM", "SWORD_2H", "SHIELD"},
    "PRIEST": {"DAGGER", "MACE_1H", "STAFF", "WAND", "OFF_HAND"},
    "ROGUE": {"AXE_1H", "DAGGER", "FIST", "MACE_1H", "SWORD_1H"},
    "SHAMAN": {"AXE_1H", "DAGGER", "FIST", "MACE_1H", "STAFF", "SHIELD"},
    "WARLOCK": {"DAGGER", "STAFF", "SWORD_1H", "WAND", "OFF_HAND"},
    "WARRIOR": {"AXE_1H", "FIST", "MACE_1H", "POLEARM", "STAFF", "SWORD_1H", "AXE_2H", "MACE_2H", "SWORD_2H", "SHIELD"},
}

ADDON_DB_SLOT_MAP = {
    "HEAD": "머리",
    "NECK": "목",
    "SHOULDERS": "어깨",
    "BACK": "망토",
    "CHEST": "가슴",
    "WRIST": "손목",
    "HANDS": "손",
    "WAIST": "허리",
    "LEGS": "다리",
    "FEET": "발",
    "RING": "반지",
    "TRINKET": "장신구",
    "WEAPON": "무기",
    "OFF_HAND": "보조장비",
    "SHIELD": "방패",
}

XLSX_SLOT_ALIASES = {
    "cape": "망토",
    "wrists": "손목",
    "bracers alternative": "손목",
    "main-hand": "무기",
    "mainhand": "무기",
    "pack leader main hand": "무기",
    "off-hand": "보조장비",
    "offhand": "보조장비",
    "pack leader off-hand": "보조장비",
    "ring 2": "반지",
    "주장비": "무기",
    "한손 무기": "무기",
    "양손 무기": "무기",
    "목걸이": "목",
    "신발": "발",
}

MPLUS_REWARD_PROFILES = {
    "mplus_end_of_dungeon": {
        "source": "mythicplus",
        "sourceLabel": "쐐기",
        "rewardContext": "end_of_dungeon",
        "rewardContextLabel": "던전 종료",
        "minKeystoneLevel": 10,
        "itemLevel": 266,
        "upgradeTrack": "Hero",
        "upgradeTrackKo": "영웅",
        "upgradeRank": "3/6",
        "displayLabel": "쐐기 영웅 트랙",
        "fullLabel": "쐐기 영웅 트랙 3/6 · 266 · 던전 종료 · M+10 이상",
        "itemString": None,
        "itemLink": None,
    },
    "mplus_great_vault_voidcore": {
        "source": "mythicplus",
        "sourceLabel": "쐐기",
        "rewardContext": "great_vault_voidcore",
        "rewardContextLabel": "위대한 금고/Voidcore",
        "minKeystoneLevel": 10,
        "itemLevel": 272,
        "upgradeTrack": "Myth",
        "upgradeTrackKo": "신화",
        "upgradeRank": "1/6",
        "displayLabel": "쐐기 신화 트랙",
        "fullLabel": "쐐기 신화 트랙 1/6 · 272 · 위대한 금고/Voidcore · M+10 이상",
        "itemString": None,
        "itemLink": None,
    },
}

SPEC_HEADER_RE = re.compile(r"^\s*\[(\d+)\]\s*=\s*\{\s*$")
OVERALL_ENTRY_RE = re.compile(
    r'\{\s*dungeon\s*=\s*(nil|"[^"]*"),\s*boss\s*=\s*(nil|"[^"]*"),\s*itemID\s*=\s*(\d+),'
    r'\s*slot\s*=\s*"([^"]+)",\s*note\s*=\s*"([^"]+)",\s*sourceType\s*=\s*"([^"]+)",'
    r'\s*sourceLabel\s*=\s*"([^"]+)"\s*\}'
)
FALLBACK_ENTRY_RE = re.compile(
    r'\{\s*dungeon\s*=\s*(nil|"[^"]*"),\s*boss\s*=\s*(nil|"[^"]*"),\s*itemID\s*=\s*(\d+),'
    r'\s*slot\s*=\s*"([^"]+)",\s*note\s*=\s*"([^"]+)"\s*\}'
)
DOC_SPEC_RE = re.compile(r"^##\s+(\d+)\)\s+(.+)$")
DOC_SLOT_RE = re.compile(r"^###\s+(.+)$")
DOC_BOLD_SOURCE_RE = re.compile(
    r"^- (?:[^*]*?)\*\*([^*]+)\*\*(?:\s*/\s*\*\*([^*]+)\*\*)?\s*[-—]\s*(.+)$"
)
DOC_TIER_LINE_RE = re.compile(r"^- ([^:]+):\s*\*\*([^*]+)\*\*(?:\s*[-—]\s*itemID:\s*\d+)?$")
COMPANION_RE = re.compile(r"^- \*\*([^*]+)\*\* \(([^)]+)\)")

DUNGEON_DATA = {
    "마법학자의 정원": {
        "en": "Magisters' Terrace",
        "aliases": ("마법학자의 정원", "magisters' terrace", "magister's terrace", "magisters terrace"),
    },
    "마이사라 동굴": {"en": "Maisara Caverns", "aliases": ("마이사라 동굴", "maisara caverns")},
    "제나스 지점": {
        "en": "Nexus-Point Xenas",
        "aliases": (
            "제나스 지점", "공결탑 제나스", "공결점 제나스",
            "nexus-point xenas", "nexus point xenas", "nexus-point",
        ),
    },
    "윈드러너 첨탑": {"en": "Windrunner Spire", "aliases": ("윈드러너 첨탑", "windrunner spire")},
    "알게타르 대학": {
        "en": "Algeth'ar Academy",
        "aliases": ("알게타르 대학", "알게타르 아카데미", "algeth'ar academy", "algethar academy"),
    },
    "삼두정의 권좌": {"en": "Seat of the Triumvirate", "aliases": ("삼두정의 권좌", "seat of the triumvirate")},
    "하늘탑": {"en": "Skyreach", "aliases": ("하늘탑", "skyreach")},
    "사론의 구덩이": {"en": "Pit of Saron", "aliases": ("사론의 구덩이", "pit of saron")},
}

RAID_SOURCE_DATA = {
    "공허 첨탑": {
        "en": "The Voidspire",
        "aliases": ("공허 첨탑", "공허첨탑", "the voidspire", "공허첨탑(the voidspire)", "공허 첨탑(the voidspire)"),
    },
    "꿈의 균열": {
        "en": "The Dreamrift",
        "aliases": ("꿈의 균열", "꿈의균열", "the dreamrift", "꿈의 균열(the dreamrift)", "꿈의균열(the dreamrift)"),
    },
    "쿠엘다나스 진격로": {
        "en": "March on Quel'Danas",
        "aliases": (
            "쿠엘다나스 진격로",
            "쿠엘다나스행군",
            "쿠엘다나스 행군",
            "march on quel'danas",
            "march on quel’danas",
            "쿠엘다나스 행군(march on quel'danas)",
            "쿠엘다나스 행군(march on quel’danas)",
        ),
    },
}

ZONE_ID_TO_DUNGEON = {
    16573: "제나스 지점",
    15808: "윈드러너 첨탑",
    14032: "알게타르 대학",
    8910: "삼두정의 권좌",
    15829: "마법학자의 정원",
    16395: "마이사라 동굴",
    6988: "하늘탑",
    4813: "사론의 구덩이",
}

SOURCE_DETAIL_KOKR = {
    "Raid": "레이드",
    "Crafting": "제작",
    "Blacksmithing": "대장기술",
    "Leatherworking": "가죽세공",
    "Tailoring": "재봉",
    "Engineering": "기계공학",
    "Jewelcrafting": "보석세공",
    "Alchemy": "연금술",
    "Enchanting": "마법부여",
    "Inscription": "주문각인",
    "Catalyst": "촉매",
    "Catalyst via": "촉매",
    "Tier Set": "티어 세트",
    "Tier": "티어",
    "Midnight Falls": "한밤 폭포",
    "March on Quel'Danas": "쿠엘다나스 진격로",
    "March on Quel’Danas": "쿠엘다나스 진격로",
    "The Voidspire": "공허 첨탑",
    "The Dreamrift": "꿈의균열",
    "Dreamrift": "꿈의균열",
    "Belo'ren": "공허 첨탑",
    "Belo’ren": "공허 첨탑",
    "Belo'ren, Child of Al'ar": "공허 첨탑",
    "Belo’ren, Child of Al’ar": "공허 첨탑",
    "Chimaerus": "꿈의균열",
    "Chimaerus the Undreamt God": "꿈의균열",
    "Chimareus, the Undreamt God": "꿈의균열",
    "Crown of the Cosmos": "공허 첨탑",
    "Fallen-King Salhadaar": "공허 첨탑",
    "Imperator Averzian": "전제군주 아베르지안",
    "Lightblinded Vanguard": "공허 첨탑",
    "L'ura": "꿈의균열",
    "L’ura": "꿈의균열",
    "Vaelgor": "바엘고어",
    "Vaelgor & Ezzorak": "바엘고어 & 에조라크",
    "Vaelgor and Ezzorak": "바엘고어 & 에조라크",
    "Vorasius": "보라시우스",
    "Alleria Windrunner": "쿠엘다나스 진격로",
    "War Chaplain Senn": "전투전도사 센",
    "Belo'ren": "벨로렌",
    "Belo’ren": "벨로렌",
    "Belo'ren, Child of Al'ar": "벨로렌",
    "Belo’ren, Child of Al’ar": "벨로렌",
    "Chimaerus": "카이메루스",
    "Chimaerus the Undreamt God": "카이메루스",
    "Chimareus, the Undreamt God": "카이메루스",
}

SLOT_HINTS = {
    "무기": "무기",
    "무기 / 보조장비": None,
    "장신구": "장신구",
    "목걸이": "목",
    "목": "목",
    "반지": "반지",
    "허리": "허리",
    "손목": "손목",
    "장갑": "손",
    "장갑(오프셋)": "손",
    "망토": "망토",
    "망토 / 신발": None,
    "가슴": "가슴",
    "다리": "다리",
    "신발": "발",
    "어깨": "어깨",
    "어깨(오프셋)": "어깨",
    "머리": "머리",
    "방패": "방패",
    "보조장비": "보조장비",
    "4세트 우선 부위": "TIER",
}

DOC_SLOT_PREFIXES = {
    "머리": "머리",
    "목": "목",
    "어깨": "어깨",
    "망토": "망토",
    "가슴": "가슴",
    "손목": "손목",
    "손": "손",
    "허리": "허리",
    "다리": "다리",
    "발": "발",
    "신발": "발",
    "장갑": "손",
    "반지": "반지",
    "장신구": "장신구",
    "무기": "무기",
    "보조무기": "무기",
    "보조장비": "보조장비",
    "방패": "방패",
}

ARMOR_SLOT_MAP = {
    1: "머리", 2: "목", 3: "어깨", 5: "가슴", 6: "허리", 7: "다리", 8: "발", 9: "손목",
    10: "손", 11: "반지", 12: "장신구", 15: "망토", 16: "망토", 23: "보조장비",
}

MANUAL_ITEM_OVERRIDES = {
    193714: {
        "slot": "손목",
        "nameEnUS": "Frenzyroot Cuffs",
        "nameKoKR": "광분뿌리 소매장식",
        "quality": 4,
        "icon": 0,
        "dungeon": "알게타르 대학",
        "boss": "",
    },
}

DOC_ITEM_NAME_ALIASES_RAW = {
    "Soulhunter's Mask": "Shroud of the Soulhunter",
    "Scabrous Zombie Leather Belt": "Scabrous Zombie Belt",
}


def normalize_quotes(text: str) -> str:
    return (
        text.replace("’", "'")
        .replace("‘", "'")
        .replace("“", '"')
        .replace("”", '"')
        .replace("–", "-")
        .replace("—", "-")
    )


def normalize_key(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", normalize_quotes(text or "")).lower()
    return re.sub(r"[^a-z0-9가-힣]+", "", normalized)


DOC_ITEM_NAME_ALIASES = {
    normalize_key(source): target
    for source, target in DOC_ITEM_NAME_ALIASES_RAW.items()
}


def clean_markdown(text: str) -> str:
    cleaned = normalize_quotes(text or "")
    cleaned = cleaned.replace("**", "")
    cleaned = re.sub(r"`([^`]*)`", r"\1", cleaned)
    cleaned = re.sub(r"\[(.*?)\]\(.*?\)", r"\1", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip(" -")


def clean_xlsx_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", normalize_quotes(str(value)).replace("\xa0", " ")).strip()


def lua_string(value: str) -> str:
    escaped = value.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def strip_lua_string(value: str) -> str:
    return "" if value == "nil" else value.strip('"')


def canonicalize_dungeon(value: str) -> Optional[str]:
    key = normalize_key(value)
    if not key:
        return None
    for dungeon_ko, info in DUNGEON_DATA.items():
        for alias in info["aliases"]:
            alias_key = normalize_key(alias)
            if alias_key == key or alias_key in key or key in alias_key:
                return dungeon_ko
    return None


def canonicalize_raid_source(value: str) -> Optional[Tuple[str, str]]:
    key = normalize_key(value)
    if not key:
        return None
    for raid_ko, info in RAID_SOURCE_DATA.items():
        for alias in info["aliases"]:
            alias_key = normalize_key(alias)
            if alias_key == key or alias_key in key or key in alias_key:
                return raid_ko, str(info["en"])
    return None


def dungeon_en_name(dungeon_ko: str) -> str:
    return str(DUNGEON_DATA[dungeon_ko]["en"])


def is_english_only(text: str) -> bool:
    return bool(re.search(r"[A-Za-z]", text or "")) and not bool(re.search(r"[가-힣]", text or ""))


def is_hangul_present(text: str) -> bool:
    return bool(re.search(r"[가-힣]", text or ""))


def parse_overall_rows(path: Path) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    text = path.read_text(encoding="utf-8")
    start = text.index("local overallOverrides = {")
    end = text.index("\n\nfor specID, overallItems in pairs(overallOverrides) do")
    block = text[start:end]

    current_spec: Optional[int] = None
    for raw_line in block.splitlines():
        spec_match = SPEC_HEADER_RE.match(raw_line)
        if spec_match:
            current_spec = int(spec_match.group(1))
            continue
        if current_spec is None:
            continue
        row_match = OVERALL_ENTRY_RE.search(raw_line)
        if not row_match:
            continue
        dungeon_raw, boss_raw, item_id_raw, slot, note, source_type, source_label = row_match.groups()
        rows.append(
            {
                "specID": current_spec,
                "dungeon": strip_lua_string(dungeon_raw),
                "boss": strip_lua_string(boss_raw),
                "itemID": int(item_id_raw),
                "slot": slot,
                "note": note,
                "sourceType": source_type,
                "sourceLabel": source_label,
                "origin": "overall",
            }
        )
    return rows


def parse_fallback_rows(path: Path) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    current_spec: Optional[int] = None
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        spec_match = SPEC_HEADER_RE.match(raw_line)
        if spec_match:
            current_spec = int(spec_match.group(1))
            continue
        if current_spec is None:
            continue
        row_match = FALLBACK_ENTRY_RE.search(raw_line)
        if not row_match:
            if raw_line.strip() == "},":
                current_spec = None
            continue
        dungeon_raw, boss_raw, item_id_raw, slot, note = row_match.groups()
        rows.append(
            {
                "specID": current_spec,
                "dungeon": strip_lua_string(dungeon_raw),
                "boss": strip_lua_string(boss_raw),
                "itemID": int(item_id_raw),
                "slot": slot,
                "note": note,
                "sourceType": "mythicplus",
                "sourceLabel": strip_lua_string(dungeon_raw),
                "origin": "fallback",
            }
        )
    return rows


def parse_companion_map(path: Path) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        match = COMPANION_RE.match(raw_line.strip())
        if not match:
            continue
        ko_name, en_name = match.groups()
        if is_hangul_present(ko_name):
            mapping[normalize_key(en_name)] = ko_name.strip()
    return mapping


def classify_doc_source(source_text: str) -> Optional[str]:
    if not source_text:
        return None
    normalized = normalize_key(source_text)
    if "제작" in source_text or normalized == normalize_key("Crafting"):
        return "crafted"
    if "촉매" in source_text or "Catalyst" in source_text or "Tier Set" in source_text:
        return "tier"
    if canonicalize_dungeon(source_text):
        return "mythicplus"
    return "raid"


def parse_doc_rows(path: Path) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    current_spec: Optional[int] = None
    current_heading: Optional[str] = None
    heading_rank = 0

    def add_doc_row(
        item_name: str,
        raw_source: str,
        slot_hint: Optional[str],
        source_group: Optional[str],
        detail_ko: Optional[str],
        detail_en: Optional[str],
    ) -> None:
        nonlocal heading_rank
        if current_spec is None or not item_name:
            return
        heading_rank += 1
        rows.append(
            {
                "specID": current_spec,
                "itemName": clean_markdown(item_name),
                "rawSource": clean_markdown(raw_source),
                "slotHint": slot_hint,
                "sourceGroupHint": source_group,
                "detailKoHint": detail_ko,
                "detailEnHint": detail_en,
                "origin": "doc",
                "docRank": heading_rank,
            }
        )

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.rstrip()
        spec_match = DOC_SPEC_RE.match(line)
        if spec_match:
            current_spec = DOC_SECTION_SPEC_IDS.get(int(spec_match.group(1)))
            current_heading = None
            heading_rank = 0
            continue

        slot_match = DOC_SLOT_RE.match(line)
        if slot_match:
            current_heading = clean_markdown(slot_match.group(1))
            if current_heading == "우선 파밍 인던":
                current_heading = None
            continue

        if current_spec is None or current_heading is None:
            continue

        tier_match = DOC_TIER_LINE_RE.match(line.strip())
        if tier_match and SLOT_HINTS.get(current_heading) == "TIER":
            slot_prefix, item_name = tier_match.groups()
            slot_hint = DOC_SLOT_PREFIXES.get(clean_markdown(slot_prefix))
            add_doc_row(item_name, "Tier Set", slot_hint, "tier", "티어 세트", "Tier Set")
            continue

        item_match = DOC_BOLD_SOURCE_RE.match(line.strip())
        if not item_match:
            continue

        first_item, second_item, raw_source = item_match.groups()
        source_text = clean_markdown(raw_source)
        source_group = classify_doc_source(source_text)
        detail_ko: Optional[str] = None
        detail_en: Optional[str] = None

        if source_group == "mythicplus":
            dungeon_ko = canonicalize_dungeon(source_text)
            if dungeon_ko:
                detail_ko = dungeon_ko
                detail_en = dungeon_en_name(dungeon_ko)
        elif source_group == "crafted":
            detail_ko = "제작"
            detail_en = "Crafting"
        elif source_group == "tier":
            detail_en = "Catalyst" if "Catalyst" in source_text else "Tier Set"
            detail_ko = SOURCE_DETAIL_KOKR.get(detail_en, "티어 세트")
        elif source_group == "raid":
            canonical_raid = canonicalize_raid_source(source_text)
            if canonical_raid:
                detail_ko, detail_en = canonical_raid
            else:
                detail_ko = SOURCE_DETAIL_KOKR.get(source_text)
                detail_en = source_text

        slot_hint = SLOT_HINTS.get(current_heading)
        if slot_hint == "TIER":
            slot_hint = None

        add_doc_row(first_item, source_text, slot_hint, source_group, detail_ko, detail_en)
        if second_item:
            add_doc_row(second_item, source_text, slot_hint, source_group, detail_ko, detail_en)

    return rows


def parse_data_page_json(text: str) -> Dict[str, object]:
    match = re.search(r'data-page="([^"]+)"', text)
    if not match:
        raise ValueError("wago.tools data-page payload missing")
    return json.loads(html.unescape(match.group(1)))


def get_latest_retail_db2_build(session: requests.Session) -> str:
    response = session.get(WAGO_DB2_INDEX_URL, timeout=REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    payload = parse_data_page_json(response.text)
    versions = payload.get("props", {}).get("versions", [])
    for version in versions:
        if isinstance(version, str) and version.startswith(RETAIL_DB2_PREFIX):
            return version
    raise ValueError(f"No DB2 build found for prefix {RETAIL_DB2_PREFIX!r}")


def fetch_wago_csv_rows(
    session: requests.Session,
    table: str,
    build: str,
    *,
    locale: Optional[str] = None,
) -> Iterable[Dict[str, str]]:
    locale_param = f"&locale={locale}" if locale else ""
    url = WAGO_DB2_URL_TEMPLATE.format(table=table, build=build, locale_param=locale_param)
    response = session.get(url, timeout=REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    return csv.DictReader(response.text.splitlines())


def resolve_slot_from_db2(class_id: int, subclass_id: int, inventory_type: int) -> Optional[str]:
    if class_id == 2:
        return "무기"
    if inventory_type == 14:
        if class_id == 4 and subclass_id == 6:
            return "방패"
        return "보조장비"
    return ARMOR_SLOT_MAP.get(inventory_type)


def load_sparse_rows(
    session: requests.Session,
    build: str,
    locale: str,
    target_ids: set[int],
    *,
    name_keys: Optional[set[str]] = None,
) -> Tuple[Dict[int, Dict[str, str]], Dict[str, List[int]]]:
    rows: Dict[int, Dict[str, str]] = {}
    candidate_ids_by_name: Dict[str, List[int]] = defaultdict(list)
    for row in fetch_wago_csv_rows(session, "ItemSparse", build, locale=locale):
        try:
            item_id = int(row.get("ID") or 0)
        except ValueError:
            continue
        display_key = normalize_key(str(row.get("Display_lang") or ""))
        if name_keys and display_key in name_keys:
            candidate_ids_by_name.setdefault(display_key, []).append(item_id)
            rows[item_id] = row
            continue
        if item_id in target_ids:
            rows[item_id] = row
            if len(rows) == len(target_ids) and not name_keys:
                break
    return rows, candidate_ids_by_name


def load_item_rows(
    session: requests.Session,
    build: str,
    target_ids: set[int],
) -> Dict[int, Dict[str, str]]:
    rows: Dict[int, Dict[str, str]] = {}
    for row in fetch_wago_csv_rows(session, "Item", build):
        try:
            item_id = int(row.get("ID") or 0)
        except ValueError:
            continue
        if item_id in target_ids:
            rows[item_id] = row
            if len(rows) == len(target_ids):
                break
    return rows


def load_icon_overrides(
    session: requests.Session,
    build: str,
    target_ids: set[int],
) -> Dict[int, int]:
    appearance_ids_by_item: Dict[int, int] = {}
    for row in fetch_wago_csv_rows(session, "ItemModifiedAppearance", build):
        try:
            item_id = int(row.get("ItemID") or 0)
            appearance_id = int(row.get("ItemAppearanceID") or 0)
        except ValueError:
            continue
        if item_id in target_ids and appearance_id and item_id not in appearance_ids_by_item:
            appearance_ids_by_item[item_id] = appearance_id
            if len(appearance_ids_by_item) == len(target_ids):
                break

    if not appearance_ids_by_item:
        return {}

    needed_appearance_ids = set(appearance_ids_by_item.values())
    icon_by_appearance: Dict[int, int] = {}
    for row in fetch_wago_csv_rows(session, "ItemAppearance", build):
        try:
            appearance_id = int(row.get("ID") or 0)
            icon_id = int(row.get("DefaultIconFileDataID") or 0)
        except ValueError:
            continue
        if appearance_id in needed_appearance_ids:
            icon_by_appearance[appearance_id] = icon_id
            if len(icon_by_appearance) == len(needed_appearance_ids):
                break

    return {
        item_id: icon_by_appearance.get(appearance_id, 0)
        for item_id, appearance_id in appearance_ids_by_item.items()
    }


def fetch_all_item_metadata(
    item_ids: Iterable[int],
    doc_name_keys: Optional[set[str]] = None,
) -> Dict[int, Dict[str, object]]:
    target_ids = sorted(set(int(item_id) for item_id in item_ids))
    target_id_set = set(target_ids)
    metadata: Dict[int, Dict[str, object]] = {}

    with requests.Session() as session:
        build = get_latest_retail_db2_build(session)
        print(f"[db2] using build {build}")

        sparse_en, candidate_ids_by_name = load_sparse_rows(
            session,
            build,
            "enUS",
            target_id_set,
            name_keys=doc_name_keys,
        )
        extra_target_ids = {
            item_id
            for ids in candidate_ids_by_name.values()
            for item_id in ids
        }
        if extra_target_ids:
            target_id_set.update(extra_target_ids)
            target_ids = sorted(target_id_set)

        sparse_ko, _ = load_sparse_rows(session, build, "koKR", target_id_set)
        item_rows = load_item_rows(session, build, target_id_set)

        missing_icon_ids = {
            item_id
            for item_id, row in item_rows.items()
            if int(row.get("IconFileDataID") or 0) == 0
        }
        icon_overrides = load_icon_overrides(session, build, missing_icon_ids)

    for item_id in target_ids:
        override = MANUAL_ITEM_OVERRIDES.get(item_id)
        en_row = sparse_en.get(item_id)
        ko_row = sparse_ko.get(item_id)
        item_row = item_rows.get(item_id)

        if override:
            metadata[item_id] = {"itemID": item_id, **override}
            continue
        if not en_row or not ko_row or not item_row:
            raise ValueError(f"DB2 metadata missing for item {item_id}")

        class_id = int(item_row.get("ClassID") or 0)
        subclass_id = int(item_row.get("SubclassID") or 0)
        inventory_type = int(item_row.get("InventoryType") or en_row.get("InventoryType") or 0)
        icon_id = int(item_row.get("IconFileDataID") or 0) or int(icon_overrides.get(item_id) or 0)

        metadata[item_id] = {
            "itemID": item_id,
            "slot": resolve_slot_from_db2(class_id, subclass_id, inventory_type),
            "nameEnUS": str(en_row.get("Display_lang") or ""),
            "nameKoKR": str(ko_row.get("Display_lang") or ""),
            "quality": int(en_row.get("OverallQualityID") or ko_row.get("OverallQualityID") or 4),
            "icon": icon_id,
            "dungeon": "",
            "boss": "",
        }
        print(f"[item] {item_id} {metadata[item_id]['nameEnUS']}")

    return metadata


def build_name_index(
    item_metadata: Dict[int, Dict[str, object]],
) -> Tuple[Dict[str, List[int]], Dict[str, str]]:
    index: Dict[str, List[int]] = {}
    en_to_ko: Dict[str, str] = {}
    for item_id, meta in item_metadata.items():
        item_key = normalize_key(str(meta["nameEnUS"]))
        if not item_key:
            continue
        index.setdefault(item_key, []).append(item_id)
        en_to_ko[item_key] = str(meta["nameKoKR"])
    return index, en_to_ko


def classify_existing_source(entry: Dict[str, object]) -> str:
    source_type = str(entry.get("sourceType") or entry.get("sourceGroup") or "")
    source_label = str(entry.get("sourceLabel") or "")
    normalized = normalize_key(source_label)

    if source_type == "tier" or normalized in {
        normalize_key("Tier Set"),
        normalize_key("Catalyst"),
        normalize_key("Catalyst via"),
        normalize_key("Matrix Catalyst"),
        normalize_key("Catalyst / Raid / Vault"),
        normalize_key("Raid | Catalyst | Vault"),
    }:
        return "tier"
    if source_type == "crafted" or normalized in {
        normalize_key("Crafting"),
        normalize_key("Blacksmithing"),
        normalize_key("Leatherworking"),
        normalize_key("Tailoring"),
        normalize_key("Engineering"),
        normalize_key("Jewelcrafting"),
        normalize_key("Alchemy"),
        normalize_key("Enchanting"),
        normalize_key("Inscription"),
    }:
        return "crafted"
    if source_type == "mythicplus" or canonicalize_dungeon(source_label or str(entry.get("dungeon") or "")):
        return "mythicplus"
    return "raid"


def normalize_xlsx_slot(slot_value: object, meta_slot: object = "") -> str:
    raw = clean_xlsx_text(slot_value)
    key = normalize_quotes(raw).lower()
    if raw in SLOT_HINTS and SLOT_HINTS[raw]:
        return str(SLOT_HINTS[raw])
    if raw in DOC_SLOT_PREFIXES:
        return DOC_SLOT_PREFIXES[raw]
    if key in XLSX_SLOT_ALIASES:
        return XLSX_SLOT_ALIASES[key]
    meta = clean_xlsx_text(meta_slot)
    if meta:
        return meta
    return raw or "기타"


def clean_raid_source_label(source_label: str) -> str:
    label = clean_xlsx_text(source_label)
    if not label:
        return ""
    label = re.sub(r"\s*\((?:Raid|레이드)\)\s*", "", label, flags=re.IGNORECASE)
    label = label.replace(" and ", " & ")
    for sep in (" in ", " - "):
        if sep in label:
            left, right = label.split(sep, 1)
            if canonicalize_raid_source(right) or canonicalize_raid_source(label):
                return left.strip()
    return label.strip(" .")


def infer_xlsx_source_group(source_type: str, source_label: str, item_id: int, legacy_lookup: Dict[int, str]) -> str:
    source_type = clean_xlsx_text(source_type)
    source_label = clean_xlsx_text(source_label)
    normalized = normalize_key(source_label)

    if source_type == "쐐기":
        return "mythicplus"
    if source_type == "제작":
        return "crafted"
    if "catalyst" in normalized or "tierset" in normalized or "티어" in normalized or normalized == normalize_key("Raid | Catalyst | Vault"):
        return "tier"
    if canonicalize_dungeon(source_label):
        return "mythicplus"
    if source_type == "레이드/티어":
        return "raid"
    if source_type in {"기타", "미지정", ""}:
        if "raid" in normalized or canonicalize_raid_source(source_label):
            return "raid"
        legacy_group = legacy_lookup.get(item_id)
        if legacy_group:
            return legacy_group
    return "raid"


def normalize_xlsx_source_label(source_group: str, source_label: str) -> str:
    label = clean_xlsx_text(source_label)
    if source_group == "tier":
        if "Catalyst" in label or "촉매" in label:
            return "Catalyst"
        return "Tier Set"
    if source_group == "crafted":
        return label or "Crafting"
    if source_group == "raid":
        return clean_raid_source_label(label) or "Raid"
    return label


def parse_xlsx_rows(path: Path) -> List[Dict[str, object]]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import the BIS xlsx source") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Long_ID"]
    header = [clean_xlsx_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: idx for idx, name in enumerate(header)}
    required = {
        "스펙키", "기준", "부위", "아이템명(한국어)", "아이템명(영어)",
        "출처", "출처유형", "Base ItemID",
    }
    missing = sorted(required - set(index))
    if missing:
        raise ValueError(f"xlsx Long_ID missing required columns: {missing}")

    rows: List[Dict[str, object]] = []
    for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in raw_row):
            continue
        spec_key = clean_xlsx_text(raw_row[index["스펙키"]])
        if spec_key not in SPEC_KEY_TO_ID:
            raise ValueError(f"Unknown xlsx spec key: {spec_key!r}")
        item_id_text = clean_xlsx_text(raw_row[index["Base ItemID"]])
        if not item_id_text:
            continue
        item_id = int(float(item_id_text))
        rows.append(
            {
                "specID": SPEC_KEY_TO_ID[spec_key],
                "specKey": spec_key,
                "criteria": clean_xlsx_text(raw_row[index["기준"]]),
                "slot": clean_xlsx_text(raw_row[index["부위"]]),
                "nameKoKR": clean_xlsx_text(raw_row[index["아이템명(한국어)"]]),
                "nameEnUS": clean_xlsx_text(raw_row[index["아이템명(영어)"]]),
                "sourceLabel": clean_xlsx_text(raw_row[index["출처"]]),
                "sourceTypeKo": clean_xlsx_text(raw_row[index["출처유형"]]),
                "itemID": item_id,
                "mplusEndLabel": clean_xlsx_text(raw_row[index.get("쐐기 +10 종료 보상 기준", -1)]) if "쐐기 +10 종료 보상 기준" in index else "",
                "mplusVaultLabel": clean_xlsx_text(raw_row[index.get("쐐기 +10 금고/Voidcore 기준", -1)]) if "쐐기 +10 금고/Voidcore 기준" in index else "",
                "origin": "xlsx",
            }
        )
    return rows


def localized_source_detail(
    source_group: str,
    source_label: str,
    dungeon_value: str,
    detail_hint_ko: Optional[str] = None,
    detail_hint_en: Optional[str] = None,
) -> Tuple[str, str, str, str]:
    if detail_hint_ko and detail_hint_en:
        if source_group == "mythicplus":
            return detail_hint_ko, detail_hint_en, detail_hint_ko, detail_hint_en
        return detail_hint_ko, detail_hint_en, "", ""
    if source_group == "mythicplus":
        dungeon_ko = canonicalize_dungeon(dungeon_value or source_label)
        if not dungeon_ko:
            raise ValueError(f"Mythic+ dungeon alias could not be normalized: {source_label!r}")
        return dungeon_ko, dungeon_en_name(dungeon_ko), dungeon_ko, dungeon_en_name(dungeon_ko)
    if source_group == "crafted":
        detail_en = source_label or "Crafting"
        return SOURCE_DETAIL_KOKR.get(detail_en, "제작"), detail_en, "", ""
    if source_group == "tier":
        detail_en = source_label or "Tier Set"
        detail_ko = "촉매" if "Catalyst" in detail_en else SOURCE_DETAIL_KOKR.get(detail_en, "티어 세트")
        return detail_ko, detail_en, "", ""
    detail_en = source_label or "Raid"
    canonical_raid = canonicalize_raid_source(detail_en)
    if canonical_raid:
        detail_ko, detail_en = canonical_raid
        return detail_ko, detail_en, "", ""
    detail_ko = SOURCE_DETAIL_KOKR.get(detail_en)
    if not detail_ko:
        raise ValueError(f"Missing Korean source localization for {detail_en!r}")
    return detail_ko, detail_en, "", ""


def resolve_doc_item_id(
    row: Dict[str, object],
    name_index: Dict[str, List[int]],
    item_metadata: Dict[int, Dict[str, object]],
    existing_entries: List[Dict[str, object]],
) -> int:
    item_key = normalize_key(str(row["itemName"]))
    item_key = normalize_key(DOC_ITEM_NAME_ALIASES.get(item_key, str(row["itemName"])))
    candidates: List[int] = []
    spec_id = int(row["specID"])

    for existing in existing_entries:
        if existing["specID"] != spec_id:
            continue
        item_id = int(existing["itemID"])
        meta = item_metadata.get(item_id)
        if meta and normalize_key(str(meta["nameEnUS"])) == item_key:
            candidates.append(item_id)

    if not candidates:
        candidates = list(name_index.get(item_key) or [])
    if not candidates:
        raise ValueError(
            f"Item ID could not be resolved from DOC seed: spec {spec_id}, item {row['itemName']!r}"
        )
    if len(candidates) == 1:
        return candidates[0]

    slot_hint = row.get("slotHint")
    if slot_hint:
        matching = [item_id for item_id in candidates if item_metadata[item_id].get("slot") == slot_hint]
        if len(matching) == 1:
            return matching[0]
        if matching:
            candidates = matching

    return sorted(candidates)[0]


def existing_row_priority(slot_seen: Dict[Tuple[int, str, str], int], entry: Dict[str, object]) -> int:
    key = (int(entry["specID"]), str(entry["slot"]), str(entry["origin"]))
    order = slot_seen.get(key, 0)
    slot_seen[key] = order + 1
    return SOURCE_WEIGHT[str(entry["origin"])] + NOTE_WEIGHT.get(str(entry.get("note") or ""), 50) + order


def doc_row_priority(slot_seen: Dict[Tuple[int, str, str], int], spec_id: int, slot: str) -> int:
    key = (spec_id, slot, "doc")
    order = slot_seen.get(key, 0)
    slot_seen[key] = order + 1
    return SOURCE_WEIGHT["doc"] + order


def build_catalog_rows(
    overall_rows: List[Dict[str, object]],
    fallback_rows: List[Dict[str, object]],
    doc_rows: List[Dict[str, object]],
    companion_map: Dict[str, str],
) -> Dict[int, List[Dict[str, object]]]:
    seed_rows = overall_rows + fallback_rows
    doc_name_keys = {normalize_key(str(row["itemName"])) for row in doc_rows if row.get("itemName")}
    item_metadata = fetch_all_item_metadata(
        (entry["itemID"] for entry in seed_rows),
        doc_name_keys=doc_name_keys,
    )
    name_index, _ = build_name_index(item_metadata)

    combined: Dict[int, List[Dict[str, object]]] = {}
    dedupe: Dict[Tuple[int, str, int, str], Dict[str, object]] = {}
    slot_seen: Dict[Tuple[int, str, str], int] = {}
    seed_rows_by_item: DefaultDict[int, List[Dict[str, object]]] = defaultdict(list)
    for entry in seed_rows:
        seed_rows_by_item[int(entry["itemID"])].append(entry)

    def add_catalog_row(row: Dict[str, object]) -> None:
        spec_id = int(row["specID"])
        dedupe_key = (spec_id, str(row["slot"]), int(row["itemID"]), str(row["sourceGroup"]))
        existing = dedupe.get(dedupe_key)
        if existing:
            if not existing.get("boss") and row.get("boss"):
                existing["boss"] = row["boss"]
            if not existing.get("dungeon") and row.get("dungeon"):
                existing["dungeon"] = row["dungeon"]
                existing["dungeonEnUS"] = row.get("dungeonEnUS", existing.get("dungeonEnUS"))
            if (
                (not existing.get("displaySourceKoKR") or existing.get("displaySourceKoKR") == "레이드")
                and row.get("displaySourceKoKR")
            ):
                existing["displaySourceKoKR"] = row["displaySourceKoKR"]
                existing["displaySourceEnUS"] = row["displaySourceEnUS"]
                existing["sourceLabel"] = row["sourceLabel"]
            existing["priority"] = min(int(existing["priority"]), int(row["priority"]))
            return
        combined.setdefault(spec_id, []).append(row)
        dedupe[dedupe_key] = row

    for entry in seed_rows:
        item_id = int(entry["itemID"])
        meta = item_metadata[item_id]
        source_group = classify_existing_source(entry)
        detail_ko, detail_en, dungeon_ko, dungeon_en = localized_source_detail(
            source_group,
            str(entry.get("sourceLabel") or ""),
            str(entry.get("dungeon") or ""),
        )
        add_catalog_row(
            {
                "specID": int(entry["specID"]),
                "slot": str(entry["slot"]),
                "itemID": item_id,
                "nameKoKR": str(meta["nameKoKR"] or companion_map.get(normalize_key(str(meta["nameEnUS"]))) or ""),
                "nameEnUS": str(meta["nameEnUS"]),
                "quality": int(meta["quality"]),
                "icon": str(meta["icon"]),
                "sourceGroup": source_group,
                "sourceType": source_group,
                "sourceLabel": detail_en,
                "displaySourceKoKR": detail_ko,
                "displaySourceEnUS": detail_en,
                "dungeon": dungeon_ko,
                "dungeonEnUS": dungeon_en,
                "boss": str(entry.get("boss") or meta.get("boss") or ""),
                "rewardProfiles": MPLUS_REWARD_PROFILES if source_group == "mythicplus" else None,
                "overallRank": 0,
                "sourceRank": 0,
                "priority": existing_row_priority(slot_seen, entry),
                "origin": str(entry["origin"]),
            }
        )

    for doc_row in doc_rows:
        item_id = resolve_doc_item_id(doc_row, name_index, item_metadata, seed_rows)
        meta = item_metadata[item_id]
        slot = str(doc_row.get("slotHint") or meta.get("slot") or "")
        if not slot:
            raise ValueError(f"Slot could not be resolved for DOC row {doc_row!r}")

        source_group = str(doc_row.get("sourceGroupHint") or "")
        raw_source = str(doc_row.get("rawSource") or "")
        detail_ko = doc_row.get("detailKoHint")
        detail_en = doc_row.get("detailEnHint")
        candidates = [
            seed
            for seed in seed_rows_by_item[item_id]
            if int(seed["specID"]) == int(doc_row["specID"])
        ] or seed_rows_by_item[item_id]
        chosen = candidates[0] if candidates else None

        if not source_group:
            if not candidates:
                raise ValueError(f"Source group could not be inferred for DOC row {doc_row!r}")
            source_group = classify_existing_source(chosen)
            raw_source = raw_source or str(chosen.get("sourceLabel") or chosen.get("dungeon") or "")

        if chosen and raw_source:
            if source_group == "mythicplus" and not canonicalize_dungeon(raw_source):
                raw_source = str(chosen.get("dungeon") or chosen.get("sourceLabel") or raw_source)
            elif source_group == "raid" and not (
                canonicalize_raid_source(raw_source) or SOURCE_DETAIL_KOKR.get(raw_source)
            ):
                raw_source = str(chosen.get("sourceLabel") or chosen.get("dungeon") or raw_source)

        detail_ko, detail_en, dungeon_ko, dungeon_en = localized_source_detail(
            source_group,
            raw_source if raw_source else str(meta.get("dungeon") or ""),
            raw_source if raw_source else str(meta.get("dungeon") or ""),
            detail_hint_ko=detail_ko,
            detail_hint_en=detail_en,
        )

        dedupe_key = (int(doc_row["specID"]), slot, item_id, source_group)
        raid_key = (int(doc_row["specID"]), slot, item_id, "raid")
        if source_group == "tier" and raid_key in dedupe:
            existing = dedupe.pop(raid_key)
            existing["sourceGroup"] = "tier"
            existing["sourceType"] = "tier"
            existing["sourceLabel"] = detail_en
            existing["displaySourceKoKR"] = detail_ko
            existing["displaySourceEnUS"] = detail_en
            existing["dungeon"] = dungeon_ko
            existing["dungeonEnUS"] = dungeon_en
            dedupe[(int(doc_row["specID"]), slot, item_id, "tier")] = existing
            continue
        if dedupe_key in dedupe:
            existing = dedupe[dedupe_key]
            if not existing.get("displaySourceKoKR") and detail_ko:
                existing["displaySourceKoKR"] = detail_ko
                existing["displaySourceEnUS"] = detail_en
                existing["dungeon"] = dungeon_ko
                existing["dungeonEnUS"] = dungeon_en
            continue

        add_catalog_row(
            {
                "specID": int(doc_row["specID"]),
                "slot": slot,
                "itemID": item_id,
                "nameKoKR": str(meta["nameKoKR"] or companion_map.get(normalize_key(str(meta["nameEnUS"]))) or ""),
                "nameEnUS": str(meta["nameEnUS"]),
                "quality": int(meta["quality"]),
                "icon": str(meta["icon"]),
                "sourceGroup": source_group,
                "sourceType": source_group,
                "sourceLabel": detail_en,
                "displaySourceKoKR": detail_ko,
                "displaySourceEnUS": detail_en,
                "dungeon": dungeon_ko,
                "dungeonEnUS": dungeon_en,
                "boss": str((chosen and chosen.get("boss")) or meta.get("boss") or ""),
                "rewardProfiles": MPLUS_REWARD_PROFILES if source_group == "mythicplus" else None,
                "overallRank": 0,
                "sourceRank": 0,
                "priority": doc_row_priority(slot_seen, int(doc_row["specID"]), slot),
                "origin": "doc",
            }
        )

    slot_sort_order = {
        "무기": 0, "보조장비": 1, "방패": 2, "머리": 3, "목": 4, "어깨": 5, "망토": 6,
        "가슴": 7, "손목": 8, "손": 9, "허리": 10, "다리": 11, "발": 12, "반지": 13, "장신구": 14,
    }
    for spec_id, rows in combined.items():
        by_slot: Dict[str, List[Dict[str, object]]] = {}
        for row in rows:
            by_slot.setdefault(str(row["slot"]), []).append(row)

        ranked_rows: List[Dict[str, object]] = []
        for slot, slot_rows in by_slot.items():
            slot_rows.sort(
                key=lambda row: (
                    int(row["priority"]),
                    SOURCE_GROUP_ORDER[str(row["sourceGroup"])],
                    str(row["displaySourceEnUS"]),
                    int(row["itemID"]),
                )
            )
            source_counts: Dict[str, int] = {}
            for overall_rank, row in enumerate(slot_rows, start=1):
                row["overallRank"] = overall_rank
                group = str(row["sourceGroup"])
                source_counts[group] = source_counts.get(group, 0) + 1
                row["sourceRank"] = source_counts[group]
                ranked_rows.append(row)

        combined[spec_id] = sorted(
            ranked_rows,
            key=lambda row: (
                slot_sort_order.get(str(row["slot"]), 999),
                int(row["overallRank"]),
                int(row["itemID"]),
            ),
        )

    return combined


def xlsx_row_priority(slot_seen: Dict[Tuple[int, str, str], int], entry: Dict[str, object]) -> int:
    criteria = str(entry.get("criteria") or "")
    if criteria == "전체 통합 BiS":
        base = 100
    elif criteria == "쐐기 기준 BiS":
        base = 200
    elif criteria == "레이드 기준 Best":
        base = 300
    elif criteria == "제작 기준 Best":
        base = 400
    else:
        base = 500
    key = (int(entry["specID"]), str(entry["slot"]), criteria)
    order = slot_seen.get(key, 0)
    slot_seen[key] = order + 1
    return base + order


def build_catalog_rows_from_xlsx(
    xlsx_rows: List[Dict[str, object]],
    legacy_rows: List[Dict[str, object]],
    companion_map: Dict[str, str],
) -> Dict[int, List[Dict[str, object]]]:
    item_metadata = fetch_all_item_metadata(entry["itemID"] for entry in xlsx_rows)
    legacy_group_by_item: Dict[int, str] = {}
    for entry in legacy_rows:
        item_id = int(entry["itemID"])
        legacy_group_by_item.setdefault(item_id, classify_existing_source(entry))

    combined: Dict[int, List[Dict[str, object]]] = {}
    dedupe: Dict[Tuple[int, str, int, str], Dict[str, object]] = {}
    slot_seen: Dict[Tuple[int, str, str], int] = {}

    def add_catalog_row(row: Dict[str, object]) -> None:
        spec_id = int(row["specID"])
        dedupe_key = (spec_id, str(row["slot"]), int(row["itemID"]), str(row["sourceGroup"]))
        existing = dedupe.get(dedupe_key)
        if existing:
            existing["priority"] = min(int(existing["priority"]), int(row["priority"]))
            if not existing.get("boss") and row.get("boss"):
                existing["boss"] = row["boss"]
            if not existing.get("rewardProfiles") and row.get("rewardProfiles"):
                existing["rewardProfiles"] = row["rewardProfiles"]
            return
        combined.setdefault(spec_id, []).append(row)
        dedupe[dedupe_key] = row

    for entry in xlsx_rows:
        item_id = int(entry["itemID"])
        meta = item_metadata[item_id]
        slot = normalize_xlsx_slot(entry.get("slot"), meta.get("slot"))
        source_group = infer_xlsx_source_group(
            str(entry.get("sourceTypeKo") or ""),
            str(entry.get("sourceLabel") or ""),
            item_id,
            legacy_group_by_item,
        )
        source_label = normalize_xlsx_source_label(source_group, str(entry.get("sourceLabel") or ""))
        try:
            detail_ko, detail_en, dungeon_ko, dungeon_en = localized_source_detail(
                source_group,
                source_label,
                source_label,
            )
        except ValueError:
            if source_group != "raid":
                raise
            detail_en = source_label or "Raid"
            detail_ko = SOURCE_DETAIL_KOKR.get(detail_en, "레이드")
            dungeon_ko, dungeon_en = "", ""

        reward_profiles = MPLUS_REWARD_PROFILES if source_group == "mythicplus" else None
        add_catalog_row(
            {
                "specID": int(entry["specID"]),
                "slot": slot,
                "itemID": item_id,
                "nameKoKR": str(entry.get("nameKoKR") or meta["nameKoKR"] or companion_map.get(normalize_key(str(meta["nameEnUS"]))) or ""),
                "nameEnUS": str(entry.get("nameEnUS") or meta["nameEnUS"]),
                "quality": int(meta["quality"]),
                "icon": str(meta["icon"]),
                "sourceGroup": source_group,
                "sourceType": source_group,
                "sourceLabel": detail_en,
                "displaySourceKoKR": detail_ko,
                "displaySourceEnUS": detail_en,
                "dungeon": dungeon_ko,
                "dungeonEnUS": dungeon_en,
                "boss": str(meta.get("boss") or ""),
                "overallRank": 0,
                "sourceRank": 0,
                "priority": xlsx_row_priority(slot_seen, entry),
                "origin": "xlsx",
                "rewardProfiles": reward_profiles,
            }
        )

    slot_sort_order = {
        "무기": 0, "보조장비": 1, "방패": 2, "머리": 3, "목": 4, "어깨": 5, "망토": 6,
        "가슴": 7, "손목": 8, "손": 9, "허리": 10, "다리": 11, "발": 12, "반지": 13, "장신구": 14,
    }
    for spec_id, rows in combined.items():
        by_slot: Dict[str, List[Dict[str, object]]] = {}
        for row in rows:
            by_slot.setdefault(str(row["slot"]), []).append(row)

        ranked_rows: List[Dict[str, object]] = []
        for slot, slot_rows in by_slot.items():
            slot_rows.sort(
                key=lambda row: (
                    int(row["priority"]),
                    SOURCE_GROUP_ORDER[str(row["sourceGroup"])],
                    str(row["displaySourceEnUS"]),
                    int(row["itemID"]),
                )
            )
            source_counts: Dict[str, int] = {}
            for overall_rank, row in enumerate(slot_rows, start=1):
                row["overallRank"] = overall_rank
                group = str(row["sourceGroup"])
                source_counts[group] = source_counts.get(group, 0) + 1
                row["sourceRank"] = source_counts[group]
                ranked_rows.append(row)

        combined[spec_id] = sorted(
            ranked_rows,
            key=lambda row: (
                slot_sort_order.get(str(row["slot"]), 999),
                int(row["overallRank"]),
                int(row["itemID"]),
            ),
        )

    return combined


def lua_unescape(value: str) -> str:
    return value.replace('\\"', '"').replace("\\\\", "\\")


def get_lua_field(line: str, field: str) -> object:
    match = re.search(rf'\b{re.escape(field)}\s*=\s*(nil|true|false|-?\d+|"(?:(?:\\.)|[^"\\])*")', line)
    if not match:
        return None
    raw = match.group(1)
    if raw == "nil":
        return None
    if raw == "true":
        return True
    if raw == "false":
        return False
    if raw.startswith('"') and raw.endswith('"'):
        return lua_unescape(raw[1:-1])
    return int(raw)


def parse_existing_catalog_rows(path: Path) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    if not path.exists():
        return rows
    current_spec: Optional[int] = None
    for line in path.read_text(encoding="utf-8").splitlines():
        spec_match = re.match(r"^\s*\[(\d+)\]\s*=\s*\{", line)
        if spec_match:
            current_spec = int(spec_match.group(1))
            continue
        if current_spec is None or "{ slot =" not in line:
            continue
        source_group = get_lua_field(line, "sourceGroup")
        if source_group not in {"raid", "crafted"}:
            continue
        rows.append(
            {
                "specID": current_spec,
                "slot": str(get_lua_field(line, "slot") or ""),
                "itemID": int(get_lua_field(line, "itemID") or 0),
                "nameKoKR": str(get_lua_field(line, "nameKoKR") or ""),
                "nameEnUS": str(get_lua_field(line, "nameEnUS") or ""),
                "quality": int(get_lua_field(line, "quality") or 4),
                "icon": get_lua_field(line, "icon"),
                "sourceGroup": str(source_group),
                "sourceType": str(get_lua_field(line, "sourceType") or source_group),
                "sourceLabel": str(get_lua_field(line, "sourceLabel") or ""),
                "displaySourceKoKR": str(get_lua_field(line, "displaySourceKoKR") or ""),
                "displaySourceEnUS": str(get_lua_field(line, "displaySourceEnUS") or ""),
                "dungeon": str(get_lua_field(line, "dungeon") or ""),
                "dungeonEnUS": str(get_lua_field(line, "dungeonEnUS") or ""),
                "boss": str(get_lua_field(line, "boss") or ""),
                "overallRank": int(get_lua_field(line, "overallRank") or 99),
                "sourceRank": int(get_lua_field(line, "sourceRank") or 99),
                "priority": 100 + int(get_lua_field(line, "overallRank") or 99)
                if source_group == "raid"
                else 420 + int(get_lua_field(line, "overallRank") or 99),
                "origin": "existing_catalog",
            }
        )
    return rows


def extract_addon_db_block(text: str, table_name: str) -> str:
    start = text.index(f"{table_name} = {{")
    end = text.find("\n-- END extracted block", start)
    if end == -1:
        end = len(text)
    return text[start:end]


def parse_quoted_field(text: str, field: str) -> str:
    match = re.search(rf'\b{re.escape(field)}\s*=\s*"([^"]*)"', text)
    return match.group(1) if match else ""


def parse_number_field(text: str, field: str) -> Optional[int]:
    match = re.search(rf'\b{re.escape(field)}\s*=\s*(\d+)', text)
    return int(match.group(1)) if match else None


def parse_bool_field(text: str, field: str) -> Optional[bool]:
    match = re.search(rf'\b{re.escape(field)}\s*=\s*(true|false)', text)
    if not match:
        return None
    return match.group(1) == "true"


def parse_addon_db_specs(text: str) -> Tuple[Dict[int, Dict[str, object]], Dict[int, Dict[str, object]]]:
    block = extract_addon_db_block(text, "SPECS")
    specs: Dict[int, Dict[str, object]] = {}
    policies: Dict[int, Dict[str, object]] = {}
    current_key: Optional[str] = None
    current_lines: List[str] = []

    def flush() -> None:
        if not current_key or not current_lines:
            return
        raw = "\n".join(current_lines)
        spec_key = ADDON_DB_SPEC_KEY_ALIASES.get(current_key, current_key)
        spec_id = SPEC_KEY_TO_ID.get(spec_key)
        if not spec_id:
            raise ValueError(f"Unknown addon DB spec key: {current_key}")
        info = {
            "specKey": current_key,
            "classKo": parse_quoted_field(raw, "classKo"),
            "specKo": parse_quoted_field(raw, "specKo"),
            "role": parse_quoted_field(raw, "role"),
            "armor": parse_quoted_field(raw, "armor"),
            "primary": parse_quoted_field(raw, "primary"),
            "tierSetId": parse_number_field(raw, "tierSetId") or 0,
            "tierSetKo": parse_quoted_field(raw, "tierSetKo"),
        }
        policy = {
            **info,
            "secondaryPriority": parse_quoted_field(raw, "secondaryPriority"),
            "statPriorityVerified": parse_bool_field(raw, "statPriorityVerified") is True,
            "statPriorityStatus": parse_quoted_field(raw, "statPriorityStatus"),
            "staticFinalBisVerified": False,
            "bisOptimalVerified": False,
            "bisStatus": parse_quoted_field(raw, "bisStatus"),
            "bisValidationLevel": "STRICT_STATIC_DB_CANNOT_CERTIFY_FINAL_BIS__RUNTIME_SIM_REQUIRED",
        }
        specs[spec_id] = info
        policies[spec_id] = policy

    for raw_line in block.splitlines():
        start_match = re.match(r"\s*([A-Z_]+)\s*=\s*\{", raw_line)
        if start_match:
            if start_match.group(1) == "SPECS":
                continue
            flush()
            current_key = start_match.group(1)
            current_lines = [raw_line]
            continue
        if current_key:
            current_lines.append(raw_line)
            if raw_line.strip() == "},":
                flush()
                current_key = None
                current_lines = []
    flush()
    return specs, policies


def parse_addon_db_dungeons(text: str) -> Dict[str, str]:
    block = extract_addon_db_block(text, "DUNGEONS")
    dungeons: Dict[str, str] = {}
    for ko, en in re.findall(r'ko\s*=\s*"([^"]+)"\s*,\s*en\s*=\s*"([^"]+)"', block):
        dungeons[ko] = en
    return dungeons


def parse_addon_db_simple_item_table(text: str, table_name: str, source_slot: Optional[str] = None) -> List[Dict[str, object]]:
    block = extract_addon_db_block(text, table_name)
    rows: List[Dict[str, object]] = []
    pattern = re.compile(
        r'\{\s*slot\s*=\s*"([^"]+)"\s*,\s*id\s*=\s*(\d+)\s*,\s*ko\s*=\s*"([^"]+)"\s*,\s*dungeon\s*=\s*"([^"]+)"\s*\}'
    )
    for slot, item_id, ko, dungeon in pattern.findall(block):
        rows.append(
            {
                "slot": ADDON_DB_SLOT_MAP.get(source_slot or slot, ADDON_DB_SLOT_MAP.get(slot, slot)),
                "itemID": int(item_id),
                "nameKoKR": ko,
                "dungeon": dungeon,
                "kind": table_name,
            }
        )
    return rows


def parse_addon_db_trinkets(text: str) -> List[Dict[str, object]]:
    block = extract_addon_db_block(text, "TRINKETS")
    rows: List[Dict[str, object]] = []
    pattern = re.compile(
        r'\{\s*id\s*=\s*(\d+)\s*,\s*ko\s*=\s*"([^"]+)"\s*,\s*usableBy\s*=\s*"([^"]+)"\s*,\s*dungeon\s*=\s*"([^"]+)"\s*\}'
    )
    for item_id, ko, usable_by, dungeon in pattern.findall(block):
        rows.append(
            {
                "slot": "장신구",
                "itemID": int(item_id),
                "nameKoKR": ko,
                "usableBy": usable_by,
                "dungeon": dungeon,
                "kind": "TRINKETS",
            }
        )
    return rows


def parse_addon_db_weapons(text: str) -> List[Dict[str, object]]:
    block = extract_addon_db_block(text, "WEAPONS")
    rows: List[Dict[str, object]] = []
    pattern = re.compile(
        r'\{\s*type\s*=\s*"([^"]+)"\s*,\s*stat\s*=\s*"([^"]+)"\s*,\s*id\s*=\s*(\d+)\s*,\s*ko\s*=\s*"([^"]+)"\s*,\s*dungeon\s*=\s*"([^"]+)"\s*\}'
    )
    for weapon_type, stat, item_id, ko, dungeon in pattern.findall(block):
        rows.append(
            {
                "slot": ADDON_DB_SLOT_MAP.get(weapon_type, "무기" if weapon_type != "SHIELD" else "방패"),
                "itemID": int(item_id),
                "nameKoKR": ko,
                "weaponType": weapon_type,
                "weaponStat": stat,
                "dungeon": dungeon,
                "kind": "WEAPONS",
            }
        )
    return rows


def parse_addon_db_tier_sets(text: str) -> Dict[int, Dict[str, object]]:
    block = extract_addon_db_block(text, "TIER_SETS")
    sets: Dict[int, Dict[str, object]] = {}
    current_set: Optional[int] = None
    current_lines: List[str] = []

    def flush() -> None:
        if current_set is None or not current_lines:
            return
        raw = "\n".join(current_lines)
        pieces = []
        for slot_key, item_id, ko, slot_ko in re.findall(
            r'([A-Z]+)\s*=\s*\{\s*id\s*=\s*(\d+)\s*,\s*ko\s*=\s*"([^"]+)"\s*,\s*slotKo\s*=\s*"([^"]+)"',
            raw,
        ):
            pieces.append(
                {
                    "slot": slot_ko,
                    "slotKey": slot_key,
                    "itemID": int(item_id),
                    "nameKoKR": ko,
                }
            )
        sets[current_set] = {
            "classKo": parse_quoted_field(raw, "classKo"),
            "setKo": parse_quoted_field(raw, "setKo"),
            "pieces": pieces,
        }

    for raw_line in block.splitlines():
        start_match = re.match(r"\s*\[(\d+)\]\s*=\s*\{", raw_line)
        if start_match:
            flush()
            current_set = int(start_match.group(1))
            current_lines = [raw_line]
            continue
        if current_set is not None:
            current_lines.append(raw_line)
            if raw_line.startswith("  },"):
                flush()
                current_set = None
                current_lines = []
    flush()
    return sets


def stat_matches_primary(stat_token: str, primary: str) -> bool:
    if stat_token == "AGILITY":
        return primary == "민첩"
    if stat_token == "STRENGTH":
        return primary == "힘"
    if stat_token == "INTELLECT":
        return primary == "지능"
    if stat_token == "AGILITY_STRENGTH":
        return primary in {"민첩", "힘"}
    if stat_token == "AGILITY_INTELLECT":
        return primary in {"민첩", "지능"}
    if stat_token == "INTELLECT_STRENGTH":
        return primary in {"지능", "힘"}
    return True


def trinket_matches_spec(row: Dict[str, object], spec: Dict[str, object]) -> bool:
    usable_by = str(row.get("usableBy") or "")
    primary = str(spec.get("primary") or "")
    role = str(spec.get("role") or "")
    if usable_by == "ALL_PRIMARY" or usable_by == "MASTERY":
        return True
    if usable_by == "AGILITY_STRENGTH":
        return primary in {"민첩", "힘"}
    if usable_by == "TANK_AGILITY_STRENGTH":
        return role == "TANK" and primary in {"민첩", "힘"}
    if usable_by == "AGILITY_INTELLECT":
        return primary in {"민첩", "지능"}
    if usable_by == "STRENGTH":
        return primary == "힘"
    if usable_by == "INTELLECT":
        return primary == "지능"
    if usable_by == "HEALER_INTELLECT":
        return role == "HEALER" and primary == "지능"
    return True


def weapon_matches_spec(row: Dict[str, object], spec_id: int, spec: Dict[str, object]) -> bool:
    class_token = SPEC_ID_TO_CLASS_TOKEN.get(spec_id)
    if class_token and row.get("weaponType") not in CLASS_WEAPON_TYPES.get(class_token, set()):
        return False
    return stat_matches_primary(str(row.get("weaponStat") or ""), str(spec.get("primary") or ""))


def addon_db_validation_meta(policy: Dict[str, object], source_group: str) -> Dict[str, object]:
    return {
        "staticFinalBisVerified": False,
        "bisValidationLevel": str(policy.get("bisValidationLevel") or "STATIC_BIS_NOT_CONFIRMED"),
        "statPriorityVerified": policy.get("statPriorityVerified") is True,
        "runtimeItemLinkRequired": source_group in {"mythicplus", "tier"},
        "mythTrackVerified": False,
        "statPrioritySummary": str(policy.get("secondaryPriority") or ""),
    }


def build_catalog_rows_from_addon_db(
    addon_db_path: Path,
    preserved_rows: List[Dict[str, object]],
) -> Tuple[Dict[int, List[Dict[str, object]]], Dict[int, Dict[str, object]]]:
    text = addon_db_path.read_text(encoding="utf-8")
    specs, spec_policies = parse_addon_db_specs(text)
    dungeon_en_by_ko = parse_addon_db_dungeons(text)
    tier_sets = parse_addon_db_tier_sets(text)
    accessory_rows = parse_addon_db_simple_item_table(text, "ACCESSORIES")
    armor_rows_by_armor = {
        "CLOTH": parse_addon_db_simple_item_table(text, "CLOTH_ARMOR"),
        "LEATHER": parse_addon_db_simple_item_table(text, "LEATHER_ARMOR"),
        "MAIL": parse_addon_db_simple_item_table(text, "MAIL_ARMOR"),
        "PLATE": parse_addon_db_simple_item_table(text, "PLATE_ARMOR"),
    }
    trinket_rows = parse_addon_db_trinkets(text)
    weapon_rows = parse_addon_db_weapons(text)

    candidate_ids = {int(row["itemID"]) for row in preserved_rows}
    for rows in armor_rows_by_armor.values():
        candidate_ids.update(int(row["itemID"]) for row in rows)
    for rows in (accessory_rows, trinket_rows, weapon_rows):
        candidate_ids.update(int(row["itemID"]) for row in rows)
    for tier in tier_sets.values():
        candidate_ids.update(int(piece["itemID"]) for piece in tier.get("pieces", []))

    item_metadata = fetch_all_item_metadata(candidate_ids)
    combined: Dict[int, List[Dict[str, object]]] = {}
    dedupe: Dict[Tuple[int, str, int, str], Dict[str, object]] = {}

    slot_sort_order = {
        "무기": 0, "보조장비": 1, "방패": 2, "머리": 3, "목": 4, "어깨": 5, "망토": 6,
        "가슴": 7, "손목": 8, "손": 9, "허리": 10, "다리": 11, "발": 12, "반지": 13, "장신구": 14,
    }

    def add_catalog_row(row: Dict[str, object]) -> None:
        spec_id = int(row["specID"])
        key = (spec_id, str(row["slot"]), int(row["itemID"]), str(row["sourceGroup"]))
        existing = dedupe.get(key)
        if existing:
            existing["priority"] = min(int(existing["priority"]), int(row["priority"]))
            return
        combined.setdefault(spec_id, []).append(row)
        dedupe[key] = row

    def make_mplus_row(spec_id: int, source: Dict[str, object], priority: int) -> Dict[str, object]:
        item_id = int(source["itemID"])
        spec_policy = spec_policies[spec_id]
        meta = item_metadata[item_id]
        dungeon_ko = str(source.get("dungeon") or "")
        dungeon_en = dungeon_en_by_ko.get(dungeon_ko) or DUNGEON_DATA.get(dungeon_ko, {}).get("en") or dungeon_ko
        return {
            "specID": spec_id,
            "slot": str(source["slot"]),
            "itemID": item_id,
            "nameKoKR": str(source.get("nameKoKR") or meta.get("nameKoKR") or ""),
            "nameEnUS": str(meta.get("nameEnUS") or source.get("nameKoKR") or ""),
            "quality": int(meta.get("quality") or 4),
            "icon": meta.get("icon"),
            "sourceGroup": "mythicplus",
            "sourceType": "mythicplus",
            "sourceLabel": dungeon_en,
            "displaySourceKoKR": dungeon_ko,
            "displaySourceEnUS": dungeon_en,
            "dungeon": dungeon_ko,
            "dungeonEnUS": dungeon_en,
            "boss": "",
            "rewardProfiles": MPLUS_REWARD_PROFILES,
            "overallRank": 0,
            "sourceRank": 0,
            "priority": priority,
            "origin": "addon_db",
            **addon_db_validation_meta(spec_policy, "mythicplus"),
        }

    def make_tier_row(spec_id: int, piece: Dict[str, object], priority: int) -> Dict[str, object]:
        item_id = int(piece["itemID"])
        spec_policy = spec_policies[spec_id]
        meta = item_metadata[item_id]
        return {
            "specID": spec_id,
            "slot": str(piece["slot"]),
            "itemID": item_id,
            "nameKoKR": str(piece.get("nameKoKR") or meta.get("nameKoKR") or ""),
            "nameEnUS": str(meta.get("nameEnUS") or piece.get("nameKoKR") or ""),
            "quality": int(meta.get("quality") or 4),
            "icon": meta.get("icon"),
            "sourceGroup": "tier",
            "sourceType": "tier",
            "sourceLabel": "Tier Set",
            "displaySourceKoKR": "티어 세트",
            "displaySourceEnUS": "Tier Set",
            "dungeon": "",
            "dungeonEnUS": "",
            "boss": "",
            "overallRank": 0,
            "sourceRank": 0,
            "priority": priority,
            "origin": "addon_db",
            **addon_db_validation_meta(spec_policy, "tier"),
        }

    for row in preserved_rows:
        add_catalog_row(dict(row))

    for spec_id, spec in specs.items():
        policy = spec_policies[spec_id]
        tier = tier_sets.get(int(spec.get("tierSetId") or 0))
        if tier:
            for index, piece in enumerate(tier.get("pieces", []), start=1):
                add_catalog_row(make_tier_row(spec_id, piece, 160 + index))

        mplus_sources: List[Tuple[int, Dict[str, object]]] = []
        for index, row in enumerate(accessory_rows, start=1):
            mplus_sources.append((220 + index, row))
        for index, row in enumerate(armor_rows_by_armor.get(str(spec.get("armor") or ""), []), start=1):
            mplus_sources.append((260 + index, row))
        for index, row in enumerate(trinket_rows, start=1):
            if trinket_matches_spec(row, spec):
                mplus_sources.append((340 + index, row))
        for index, row in enumerate(weapon_rows, start=1):
            if weapon_matches_spec(row, spec_id, spec):
                mplus_sources.append((430 + index, row))

        for priority, source in mplus_sources:
            add_catalog_row(make_mplus_row(spec_id, source, priority))

    for spec_id, rows in combined.items():
        by_slot: Dict[str, List[Dict[str, object]]] = {}
        for row in rows:
            by_slot.setdefault(str(row["slot"]), []).append(row)
        ranked_rows: List[Dict[str, object]] = []
        for slot, slot_rows in by_slot.items():
            slot_rows.sort(
                key=lambda row: (
                    int(row["priority"]),
                    SOURCE_GROUP_ORDER[str(row["sourceGroup"])],
                    str(row.get("displaySourceEnUS") or ""),
                    int(row["itemID"]),
                )
            )
            source_counts: Dict[str, int] = {}
            for overall_rank, row in enumerate(slot_rows, start=1):
                row["overallRank"] = overall_rank
                source_group = str(row["sourceGroup"])
                source_counts[source_group] = source_counts.get(source_group, 0) + 1
                row["sourceRank"] = source_counts[source_group]
                ranked_rows.append(row)
        combined[spec_id] = sorted(
            ranked_rows,
            key=lambda row: (
                slot_sort_order.get(str(row["slot"]), 999),
                int(row["overallRank"]),
                int(row["itemID"]),
            ),
        )

    return combined, spec_policies


def render_entry(entry: Dict[str, object]) -> str:
    def optional_string(value: str) -> str:
        return "nil" if not value else lua_string(value)

    def optional_icon(value: object) -> str:
        try:
            icon_id = int(value or 0)
        except (TypeError, ValueError):
            icon_id = 0
        return "nil" if icon_id <= 0 else str(icon_id)

    def lua_value(value: object) -> str:
        if value is None:
            return "nil"
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        return lua_string(str(value))

    def render_profiles(profiles: object) -> str:
        if not isinstance(profiles, dict) or not profiles:
            return ""
        parts: List[str] = []
        fields = [
            "source",
            "sourceLabel",
            "rewardContext",
            "rewardContextLabel",
            "minKeystoneLevel",
            "itemLevel",
            "upgradeTrack",
            "upgradeTrackKo",
            "upgradeRank",
            "displayLabel",
            "fullLabel",
            "itemString",
            "itemLink",
        ]
        for profile_key in sorted(profiles):
            profile = profiles[profile_key]
            if not isinstance(profile, dict):
                continue
            profile_body = ", ".join(
                f"{field} = {lua_value(profile.get(field))}"
                for field in fields
                if field in profile
            )
            parts.append(f"{profile_key} = {{ {profile_body} }}")
        return "rewardProfiles = { " + ", ".join(parts) + " }, " if parts else ""

    def render_extra_fields(entry: Dict[str, object]) -> str:
        fields = [
            "staticFinalBisVerified",
            "bisValidationLevel",
            "statPriorityVerified",
            "runtimeItemLinkRequired",
            "mythTrackVerified",
            "statPrioritySummary",
        ]
        parts = [
            f"{field} = {lua_value(entry[field])}"
            for field in fields
            if field in entry
        ]
        return ", ".join(parts) + ", " if parts else ""

    return (
        "        { "
        f"slot = {lua_string(str(entry['slot']))}, "
        f"itemID = {int(entry['itemID'])}, "
        f"nameKoKR = {lua_string(str(entry['nameKoKR']))}, "
        f"nameEnUS = {lua_string(str(entry['nameEnUS']))}, "
        f"quality = {int(entry['quality'])}, "
        f"icon = {optional_icon(entry.get('icon'))}, "
        f"sourceGroup = {lua_string(str(entry['sourceGroup']))}, "
        f"sourceType = {lua_string(str(entry['sourceType']))}, "
        f"sourceLabel = {lua_string(str(entry['sourceLabel']))}, "
        f"displaySourceKoKR = {lua_string(str(entry['displaySourceKoKR']))}, "
        f"displaySourceEnUS = {lua_string(str(entry['displaySourceEnUS']))}, "
        f"dungeon = {optional_string(str(entry.get('dungeon') or ''))}, "
        f"dungeonEnUS = {optional_string(str(entry.get('dungeonEnUS') or ''))}, "
        f"boss = {optional_string(str(entry.get('boss') or ''))}, "
        f"{render_profiles(entry.get('rewardProfiles'))}"
        f"{render_extra_fields(entry)}"
        f"overallRank = {int(entry['overallRank'])}, "
        f"sourceRank = {int(entry['sourceRank'])} "
        "},"
    )


def render_spec_policy(spec_id: int, policy: Dict[str, object]) -> str:
    fields = [
        "specKey",
        "classKo",
        "specKo",
        "role",
        "armor",
        "primary",
        "tierSetId",
        "tierSetKo",
        "secondaryPriority",
        "statPriorityVerified",
        "statPriorityStatus",
        "staticFinalBisVerified",
        "bisOptimalVerified",
        "bisStatus",
        "bisValidationLevel",
    ]

    def lua_value(value: object) -> str:
        if value is None:
            return "nil"
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        return lua_string(str(value))

    body = ", ".join(
        f"{field} = {lua_value(policy.get(field))}"
        for field in fields
        if field in policy
    )
    return f"    [{spec_id}] = {{ {body} }},"


def render_catalog(
    catalog: Dict[int, List[Dict[str, object]]],
    spec_policies: Optional[Dict[int, Dict[str, object]]] = None,
) -> str:
    lines = [
        "local _, ns = ...",
        "",
        "-- Generated by scripts/build_bis_catalog.py",
        "-- Unified Midnight Season 1 BIS catalog.",
        "ns.Data = ns.Data or {}",
        "ns.Data.BISItems = {",
        "",
    ]
    for spec_id in sorted(catalog):
        lines.append(f"    [{spec_id}] = {{")
        for entry in catalog[spec_id]:
            lines.append(render_entry(entry))
        lines.append("    },")
        lines.append("")
    lines.append("}")
    lines.append("")
    if spec_policies:
        lines.extend(
            [
                "ns.Data.BISSpecPolicies = {",
            ]
        )
        for spec_id in sorted(spec_policies):
            lines.append(render_spec_policy(spec_id, spec_policies[spec_id]))
        lines.append("}")
        lines.append("")
    return "\n".join(lines)


def validate_catalog(catalog: Dict[int, List[Dict[str, object]]]) -> None:
    if len(catalog) != len(SPEC_NAMES):
        raise ValueError(f"Expected {len(SPEC_NAMES)} specs, got {len(catalog)}")
    missing_specs = sorted(set(SPEC_NAMES) - set(catalog))
    if missing_specs:
        raise ValueError(f"Missing specs in catalog: {missing_specs}")

    for spec_id, rows in catalog.items():
        if not rows:
            raise ValueError(f"Spec {spec_id} has no BIS rows")
        slot_counts: Dict[str, int] = {}
        source_rank_state: Dict[Tuple[str, str], int] = {}
        for row in rows:
            if not int(row["itemID"]):
                raise ValueError(f"Unresolved item ID in spec {spec_id}")
            if str(row["sourceGroup"]) not in SOURCE_GROUP_ORDER:
                raise ValueError(f"Invalid sourceGroup in spec {spec_id}: {row['sourceGroup']!r}")
            if str(row["sourceGroup"]) == "mythicplus":
                profiles = row.get("rewardProfiles")
                if not isinstance(profiles, dict):
                    raise ValueError(f"Mythic+ row missing rewardProfiles in spec {spec_id}, item {row['itemID']}")
                end_profile = profiles.get("mplus_end_of_dungeon")
                vault_profile = profiles.get("mplus_great_vault_voidcore")
                if not isinstance(end_profile, dict) or not isinstance(vault_profile, dict):
                    raise ValueError(f"Mythic+ row has incomplete rewardProfiles in spec {spec_id}, item {row['itemID']}")
                if end_profile.get("upgradeTrack") != "Hero" or int(end_profile.get("itemLevel") or 0) != 266:
                    raise ValueError(f"Invalid M+ end reward profile in spec {spec_id}, item {row['itemID']}")
                if vault_profile.get("upgradeTrack") != "Myth" or int(vault_profile.get("itemLevel") or 0) != 272:
                    raise ValueError(f"Invalid M+ vault reward profile in spec {spec_id}, item {row['itemID']}")
                for profile_key, profile in profiles.items():
                    if not isinstance(profile, dict):
                        raise ValueError(
                            f"Mythic+ row has invalid reward profile in spec {spec_id}, item {row['itemID']}: {profile_key}"
                        )
                    if profile.get("itemString") or profile.get("itemLink"):
                        raise ValueError(
                            f"Mythic+ row must not contain static itemString/itemLink in spec {spec_id}, "
                            f"item {row['itemID']}, profile {profile_key}"
                        )
                if row.get("runtimeItemLinkRequired") is not True:
                    raise ValueError(f"Mythic+ row missing runtime link policy in spec {spec_id}, item {row['itemID']}")
            if not str(row["nameKoKR"]) or is_english_only(str(row["nameKoKR"])):
                raise ValueError(
                    f"koKR item name leak in spec {spec_id}, item {row['itemID']}: {row['nameKoKR']!r}"
                )
            if is_hangul_present(str(row["nameEnUS"])):
                raise ValueError(
                    f"enUS item name leak in spec {spec_id}, item {row['itemID']}: {row['nameEnUS']!r}"
                )
            if str(row["displaySourceKoKR"]) and is_english_only(str(row["displaySourceKoKR"])):
                raise ValueError(
                    f"koKR source leak in spec {spec_id}, item {row['itemID']}: {row['displaySourceKoKR']!r}"
                )
            slot_key = str(row["slot"])
            slot_counts[slot_key] = slot_counts.get(slot_key, 0) + 1
            if int(row["overallRank"]) != slot_counts[slot_key]:
                raise ValueError(
                    f"Invalid overallRank for spec {spec_id}, item {row['itemID']}: "
                    f"expected {slot_counts[slot_key]}, got {row['overallRank']}"
                )
            key = (slot_key, str(row["sourceGroup"]))
            expected = source_rank_state.get(key, 0) + 1
            source_rank_state[key] = expected
            if int(row["sourceRank"]) != expected:
                raise ValueError(
                    f"Invalid sourceRank for spec {spec_id}, item {row['itemID']}: "
                    f"expected {expected}, got {row['sourceRank']}"
                )


def main() -> int:
    parser = argparse.ArgumentParser(description="Build ABProfileManager BIS catalog.")
    parser.add_argument(
        "--xlsx",
        type=Path,
        help="Optional Midnight S1 12.0.5 BIS xlsx source. Uses Long_ID as the catalog source.",
    )
    parser.add_argument(
        "--addon-db",
        nargs="?",
        const=ADDON_DB_FILE,
        type=Path,
        help="Optional Midnight S1 addon Lua DB source. Defaults to DOC/MidnightS1_MPlus_Addon_DB_v1.0.lua.",
    )
    args = parser.parse_args()

    spec_policies = None
    if args.addon_db:
        preserved_rows = parse_existing_catalog_rows(TARGET_FILE)
        catalog, spec_policies = build_catalog_rows_from_addon_db(args.addon_db, preserved_rows)
    else:
        overall_rows = parse_overall_rows(OVERALL_FILE)
        fallback_rows = parse_fallback_rows(FALLBACK_FILE)
        companion_map = parse_companion_map(DOC_COMPANION_FILE)
    if args.addon_db:
        pass
    elif args.xlsx:
        xlsx_rows = parse_xlsx_rows(args.xlsx)
        catalog = build_catalog_rows_from_xlsx(xlsx_rows, overall_rows + fallback_rows, companion_map)
    else:
        doc_rows = parse_doc_rows(DOC_FINAL_FILE)
        catalog = build_catalog_rows(overall_rows, fallback_rows, doc_rows, companion_map)
    validate_catalog(catalog)
    TARGET_FILE.write_text(render_catalog(catalog, spec_policies), encoding="utf-8")
    print(f"Updated {TARGET_FILE}")
    print(f"Specs: {len(catalog)}")
    print(f"Rows: {sum(len(rows) for rows in catalog.values())}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
