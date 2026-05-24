#!/usr/bin/env python3
"""Validate generated BIS reward profiles against the Midnight S1 xlsx source."""

from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, Set, Tuple


SPEC_KEY_TO_ID = {
    "MAGE_ARCANE": 62,
    "MAGE_FIRE": 63,
    "MAGE_FROST": 64,
    "PALADIN_HOLY": 65,
    "PALADIN_PROTECTION": 66,
    "PALADIN_RETRIBUTION": 70,
    "WARRIOR_ARMS": 71,
    "WARRIOR_FURY": 72,
    "WARRIOR_PROTECTION": 73,
    "DRUID_BALANCE": 102,
    "DRUID_FERAL": 103,
    "DRUID_GUARDIAN": 104,
    "DRUID_RESTORATION": 105,
    "DEATHKNIGHT_BLOOD": 250,
    "DEATHKNIGHT_FROST": 251,
    "DEATHKNIGHT_UNHOLY": 252,
    "HUNTER_BEAST_MASTERY": 253,
    "HUNTER_MARKSMANSHIP": 254,
    "HUNTER_SURVIVAL": 255,
    "PRIEST_DISCIPLINE": 256,
    "PRIEST_HOLY": 257,
    "PRIEST_SHADOW": 258,
    "ROGUE_ASSASSINATION": 259,
    "ROGUE_OUTLAW": 260,
    "ROGUE_SUBTLETY": 261,
    "SHAMAN_ELEMENTAL": 262,
    "SHAMAN_ENHANCEMENT": 263,
    "SHAMAN_RESTORATION": 264,
    "WARLOCK_AFFLICTION": 265,
    "WARLOCK_DEMONOLOGY": 266,
    "WARLOCK_DESTRUCTION": 267,
    "MONK_BREWMASTER": 268,
    "MONK_WINDWALKER": 269,
    "MONK_MISTWEAVER": 270,
    "DEMONHUNTER_HAVOC": 577,
    "DEMONHUNTER_VENGEANCE": 581,
    "DEMONHUNTER_DEVOURER": 1382,
    "EVOKER_DEVASTATION": 1467,
    "EVOKER_PRESERVATION": 1468,
    "EVOKER_AUGMENTATION": 1473,
}

DUNGEON_ALIASES = (
    "magisters terrace",
    "magister's terrace",
    "maisara caverns",
    "nexus point xenas",
    "nexus-point xenas",
    "windrunner spire",
    "algethar academy",
    "algeth'ar academy",
    "seat of the triumvirate",
    "skyreach",
    "pit of saron",
    "마법학자의 정원",
    "마이사라 동굴",
    "제나스 지점",
    "윈드러너 첨탑",
    "알게타르 대학",
    "삼두정의 권좌",
    "하늘탑",
    "사론의 구덩이",
)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize_key(text: str) -> str:
    text = unicodedata.normalize("NFKC", text or "").lower()
    text = text.replace("’", "'").replace("‘", "'")
    return re.sub(r"[^a-z0-9가-힣]+", "", text)


def is_mplus_source(source_type: str, source_label: str) -> bool:
    if source_type == "쐐기":
        return True
    source_key = normalize_key(source_label)
    return any(normalize_key(alias) in source_key for alias in DUNGEON_ALIASES)


def load_xlsx_mplus_keys(path: Path) -> Set[Tuple[int, int]]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to validate the BIS xlsx source") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Long_ID"]
    header = [clean_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: idx for idx, name in enumerate(header)}
    keys: Set[Tuple[int, int]] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        spec_key = clean_text(row[index["스펙키"]])
        if spec_key not in SPEC_KEY_TO_ID:
            raise ValueError(f"Unknown spec key in xlsx: {spec_key!r}")
        source_type = clean_text(row[index["출처유형"]])
        source_label = clean_text(row[index["출처"]])
        if not is_mplus_source(source_type, source_label):
            continue
        item_id = int(float(clean_text(row[index["Base ItemID"]])))
        keys.add((SPEC_KEY_TO_ID[spec_key], item_id))
    return keys


def parse_catalog(path: Path) -> Tuple[Set[int], Set[Tuple[int, int]], Dict[Tuple[int, int], str]]:
    spec_ids: Set[int] = set()
    mplus_keys: Set[Tuple[int, int]] = set()
    mplus_lines: Dict[Tuple[int, int], str] = {}
    current_spec = 0
    spec_re = re.compile(r"^\s*\[(\d+)\]\s*=\s*\{")
    item_re = re.compile(r"itemID\s*=\s*(\d+)")
    source_re = re.compile(r'sourceGroup\s*=\s*"([^"]+)"')
    for line in path.read_text(encoding="utf-8").splitlines():
        spec_match = spec_re.search(line)
        if spec_match:
            current_spec = int(spec_match.group(1))
            spec_ids.add(current_spec)
            continue
        if "{ slot =" not in line:
            continue
        item_match = item_re.search(line)
        source_match = source_re.search(line)
        if not item_match or not source_match:
            continue
        if source_match.group(1) != "mythicplus":
            continue
        key = (current_spec, int(item_match.group(1)))
        mplus_keys.add(key)
        mplus_lines[key] = line
    return spec_ids, mplus_keys, mplus_lines


def require_profile(line: str, token: str) -> None:
    if token not in line:
        raise ValueError(f"Catalog Mythic+ reward profile missing token {token!r}: {line[:240]}")


def validate_profiles(lines: Iterable[str]) -> None:
    for line in lines:
        for token in (
            "rewardProfiles",
            "mplus_end_of_dungeon",
            'rewardContext = "end_of_dungeon"',
            "itemLevel = 266",
            'upgradeTrack = "Hero"',
            'upgradeRank = "3/6"',
            "mplus_great_vault_voidcore",
            'rewardContext = "great_vault_voidcore"',
            "itemLevel = 272",
            'upgradeTrack = "Myth"',
            'upgradeRank = "1/6"',
        ):
            require_profile(line, token)


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate BIS reward profiles generated from the Midnight S1 xlsx source.")
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--catalog", default=Path("ABProfileManager/Data/BISCatalog.lua"), type=Path)
    args = parser.parse_args()

    xlsx_mplus_keys = load_xlsx_mplus_keys(args.xlsx)
    spec_ids, catalog_mplus_keys, catalog_mplus_lines = parse_catalog(args.catalog)

    if len(spec_ids) != 40:
        raise ValueError(f"Expected 40 specs in catalog, found {len(spec_ids)}")
    missing = sorted(xlsx_mplus_keys - catalog_mplus_keys)
    if missing:
        preview = ", ".join(f"{spec}:{item}" for spec, item in missing[:20])
        raise ValueError(f"Catalog missing {len(missing)} xlsx Mythic+ spec/item rows: {preview}")
    validate_profiles(catalog_mplus_lines.values())

    print("ok")
    print(f"specs={len(spec_ids)}")
    print(f"xlsx_mplus_unique={len(xlsx_mplus_keys)}")
    print(f"catalog_mplus_unique={len(catalog_mplus_keys)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
